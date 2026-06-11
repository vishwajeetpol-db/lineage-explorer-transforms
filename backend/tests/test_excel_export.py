"""
Excel export — dependency layering (pure) and full workbook assembly.

The layering test is pure Python and always runs. The workbook test needs
openpyxl (installed via requirements in CI) and is skipped if it's absent.
"""
import types
import pytest

from backend.excel_export import _layer_nodes, build_lineage_workbook
from backend.models import TableNode, EntityNode, LineageEdge, LineageResponse


def _tn(fqdn, name, status="connected", up=0, down=0):
    return TableNode(id=fqdn, name=name, full_name=fqdn, table_type="MANAGED",
                     lineage_status=status, upstream_count=up, downstream_count=down)


def test_layering_medallion_flow():
    ids = {
        "c.s.raw": types.SimpleNamespace(full_name="c.s.raw", name="raw"),
        "c.s.clean": types.SimpleNamespace(full_name="c.s.clean", name="clean"),
        "c.s.gold": types.SimpleNamespace(full_name="c.s.gold", name="gold"),
        "c.s.orphan": types.SimpleNamespace(full_name="c.s.orphan", name="orphan"),
    }
    edges = [("c.s.raw", "c.s.clean"), ("c.s.clean", "c.s.gold")]
    layers, orphans, adj_down = _layer_nodes(ids, edges)
    assert [ids[f].name for f in layers[0]] == ["raw"]
    assert [ids[f].name for f in layers[1]] == ["clean"]
    assert [ids[f].name for f in layers[2]] == ["gold"]
    assert [ids[f].name for f in orphans] == ["orphan"]
    assert adj_down["c.s.raw"] == {"c.s.clean"}


def test_layering_longest_path_with_join():
    """A node fed by two layers lands at the deeper layer (longest path)."""
    ids = {f: types.SimpleNamespace(full_name=f, name=f) for f in ["a", "b", "c", "d"]}
    # a->b->d and a->d : d should be at layer 2 (via b), not 1
    layers, _, _ = _layer_nodes(ids, [("a", "b"), ("b", "d"), ("a", "d")])
    layer_of = {f: L for L, fs in layers.items() for f in fs}
    assert layer_of["d"] == 2


def test_layering_cycle_safe():
    ids = {f: types.SimpleNamespace(full_name=f, name=f) for f in ["a", "b"]}
    layers, orphans, _ = _layer_nodes(ids, [("a", "b"), ("b", "a")])
    # No infinite loop; both nodes placed somewhere, none dropped.
    placed = [f for fs in layers.values() for f in fs]
    assert set(placed) == {"a", "b"}


def test_layering_fanout_job_does_not_collapse_downstream():
    """Regression: a job that both reads and writes overlapping tables (the
    'Pipeline Demo' case) used to create cycles that dumped deep nodes into
    layer 0. Deep nodes must now get real layers via cycle-robust layering.

    Graph: raw -> jobA -> mid -> jobB -> deep, where jobA also reads `mid`
    (read+write overlap → a cycle through jobA).
    """
    names = ["raw", "mid", "deep", "entity:JOB:A", "entity:JOB:B"]
    ids = {n: types.SimpleNamespace(name=n) for n in names}
    edges = [
        ("raw", "entity:JOB:A"),
        ("entity:JOB:A", "mid"),
        ("mid", "entity:JOB:A"),       # overlap → back-edge / cycle
        ("mid", "entity:JOB:B"),
        ("entity:JOB:B", "deep"),
    ]
    layers, orphans, _ = _layer_nodes(ids, edges)
    layer_of = {n: L for L, ns in layers.items() for n in ns}
    assert layer_of["raw"] == 0
    assert layer_of["deep"] > layer_of["mid"] > layer_of["raw"]  # deep is NOT at layer 0
    assert not orphans


def test_build_workbook_has_all_sheets_including_map():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from io import BytesIO

    nodes = [
        _tn("c.s.raw", "raw", "root", up=0, down=1),
        _tn("c.s.clean", "clean", "connected", up=1, down=1),
        _tn("c.s.gold", "gold", "leaf", up=1, down=0),
        _tn("c.s.lonely", "lonely", "orphan"),
        EntityNode(id="entity:PIPELINE:p1", entity_type="PIPELINE", entity_id="p1", cost_usd=12.5),
    ]
    edges = [
        LineageEdge(source="c.s.raw", target="entity:PIPELINE:p1"),
        LineageEdge(source="entity:PIPELINE:p1", target="c.s.clean"),
        LineageEdge(source="c.s.clean", target="c.s.gold"),
    ]
    result = LineageResponse(nodes=nodes, edges=edges)
    # Real recorded pairs are passed in explicitly (not derived by cross-product).
    table_edges = [{"source": "c.s.raw", "target": "c.s.clean",
                    "entity_type": "PIPELINE", "entity_id": "p1"}]
    entity_names = {"entity:PIPELINE:p1": "Clean & Ingest"}
    data = build_lineage_workbook("c", "s", result, None, entity_names, table_edges)
    assert data[:2] == b"PK"  # zip signature

    wb = load_workbook(BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Tables" in wb.sheetnames
    assert "Lineage" in wb.sheetnames
    assert "Pipelines" in wb.sheetnames
    assert "Lineage Map" in wb.sheetnames  # single-schema scope → one map sheet

    # Lineage sheet: Source | Via (pipeline) | Target — using the real pair.
    rows = [(r[0].value, r[1].value, r[2].value) for r in wb["Lineage"].iter_rows(min_row=2)]
    assert ("c.s.raw", "Clean & Ingest", "c.s.clean") in rows
