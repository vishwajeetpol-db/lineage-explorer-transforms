"""
Server-side Excel export — builds a styled multi-sheet workbook from a
LineageResponse using openpyxl (a mature OOXML library, so we don't hand-roll
the format). openpyxl is imported lazily inside build_lineage_workbook so the
rest of the app keeps working even if the dependency is somehow missing — the
export endpoint surfaces a clean 503 in that case.

Sheets: Summary, Tables, Lineage (table -> table), Pipelines, Column Lineage.
"""
from __future__ import annotations

import logging
from collections import defaultdict, deque
from datetime import datetime, timezone
from io import BytesIO

logger = logging.getLogger(__name__)


# Palette mirrors the app's dark-UI accent colors (used as cell styling).
_INDIGO = "4F46E5"
_STATUS_FILL = {
    "orphan": "FEF3C7",     # amber
    "root": "E0F2FE",       # sky
    "leaf": "EDE9FE",       # violet
    "connected": "D1FAE5",  # emerald
}
_STATUS_FONT = {
    "orphan": "92400E",
    "root": "075985",
    "leaf": "5B21B6",
    "connected": "065F46",
}


def _is_entity(node_id: str) -> bool:
    return node_id.startswith("entity:")


def _collapse_edges(edges) -> list[tuple[str, str]]:
    """Collapse entity-mediated edges (table -> pipeline -> table) into direct
    table -> table edges; pure table edges pass through."""
    entity_sources: dict[str, set[str]] = {}
    entity_targets: dict[str, set[str]] = {}
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str]] = []

    def add(s: str, t: str):
        if s == t or (s, t) in seen:
            return
        seen.add((s, t))
        out.append((s, t))

    for e in edges:
        se, te = _is_entity(e.source), _is_entity(e.target)
        if not se and not te:
            add(e.source, e.target)
        elif not se and te:
            entity_sources.setdefault(e.target, set()).add(e.source)
        elif se and not te:
            entity_targets.setdefault(e.source, set()).add(e.target)

    for entity, sources in entity_sources.items():
        for t in entity_targets.get(entity, ()):
            for s in sources:
                add(s, t)
    return out


def _split_fqdn(full_name: str) -> tuple[str, str]:
    parts = full_name.split(".")
    return (parts[0], parts[1]) if len(parts) == 3 else ("", "")


# Box geometry for the Lineage Map sheet (in cells).
_NODE_W, _NODE_H, _GAP_COL, _GAP_ROW = 4, 3, 1, 2
_TITLE_ROW, _LAYER_HDR_ROW, _FIRST_NODE_ROW = 1, 2, 4


def _layer_nodes(ids, edges):
    """Assign each connected node a dependency depth (longest path from a source).

    Works on the FULL graph (tables AND pipeline/job entities) using the raw
    edges — we no longer collapse entities into a table→table cross-product,
    because that explodes fan-out jobs into dense cyclic graphs.

    Cycle-robust: back-edges are removed via DFS to form a DAG first, then
    longest-path layering (Kahn) runs on the DAG. Without this, nodes inside a
    cycle never get a layer and pile up at layer 0.

    Pure function (no openpyxl) so the layout logic is unit-testable.
    Returns (layers_map, orphans, adj_down):
    - layers_map: {layer_index: [node_id, ...]} sorted by node name/id
    - orphans: nodes with no incident edge at all
    - adj_down: {node_id: set(downstream node_ids)} — for flow arrows
    """
    adj: dict[str, set[str]] = {n: set() for n in ids}
    incident: set[str] = set()
    for s, t in edges:
        if s in ids and t in ids and s != t:
            adj[s].add(t)
            incident.add(s)
            incident.add(t)

    orphans = [n for n in ids if n not in incident]
    connected = [n for n in ids if n in incident]

    # Strip back-edges via iterative DFS (Sugiyama cycle removal) → DAG `forward`.
    WHITE, GRAY, BLACK = 0, 1, 2
    color = {n: WHITE for n in connected}
    forward: dict[str, set[str]] = {n: set() for n in connected}
    for root in connected:
        if color[root] != WHITE:
            continue
        color[root] = GRAY
        stack = [(root, iter(sorted(adj[root])))]
        while stack:
            u, it = stack[-1]
            descended = False
            for v in it:
                if v not in color:
                    continue
                if color[v] == GRAY:
                    continue  # back-edge → drop for layering
                forward[u].add(v)
                if color[v] == WHITE:
                    color[v] = GRAY
                    stack.append((v, iter(sorted(adj[v]))))
                    descended = True
                    break
            if not descended:
                color[u] = BLACK
                stack.pop()

    # Longest-path layering on the DAG.
    indeg = {n: 0 for n in connected}
    for u in connected:
        for v in forward[u]:
            indeg[v] += 1
    layer = {n: 0 for n in connected}
    q = deque([n for n in connected if indeg[n] == 0])
    while q:
        u = q.popleft()
        for v in forward[u]:
            if layer[u] + 1 > layer[v]:
                layer[v] = layer[u] + 1
            indeg[v] -= 1
            if indeg[v] == 0:
                q.append(v)

    layers_map: dict[int, list[str]] = defaultdict(list)
    name_of = lambda n: (getattr(ids[n], "name", None) or n).lower()
    for n in connected:
        layers_map[layer[n]].append(n)
    for nodes in layers_map.values():
        nodes.sort(key=name_of)
    return layers_map, orphans, adj


_ENTITY_FILL = "E0E7FF"   # light indigo for job/pipeline boxes
_ENTITY_FONT = "3730A3"
_MAX_SCHEMA_MAPS = 12     # cap map sheets so a many-schema catalog stays sane


def _schema_of(full_name: str) -> str:
    parts = full_name.split(".")
    return parts[1] if len(parts) == 3 else ""


def _add_per_schema_maps(wb, all_nodes, raw_edges, entity_names) -> None:
    """For catalog scope: one lineage-map sheet per schema (largest first),
    instead of a single dozens-of-columns-wide grid that interleaves unrelated
    schemas. Each map's subgraph = the schema's tables + the pipelines that
    touch them + every table those pipelines read/write. Schemas with no flow
    are skipped; capped at _MAX_SCHEMA_MAPS."""
    id_node = {n.id: n for n in all_nodes}
    is_entity = lambda i: i.startswith("entity:")

    adj = defaultdict(set)  # undirected, for subgraph expansion
    for s, t in raw_edges:
        if s in id_node and t in id_node:
            adj[s].add(t)
            adj[t].add(s)

    by_schema = defaultdict(list)
    for n in all_nodes:
        if getattr(n, "node_type", None) == "table":
            by_schema[_schema_of(n.full_name)].append(n.id)

    used_names: set[str] = set()
    made = 0
    for sch in sorted(by_schema, key=lambda s: -len(by_schema[s])):
        if made >= _MAX_SCHEMA_MAPS:
            break
        tids = by_schema[sch]
        include = set(tids)
        # pipelines touching this schema's tables + all tables they touch
        for tid in tids:
            for nb in adj.get(tid, ()):
                include.add(nb)
                if is_entity(nb):
                    include.update(adj.get(nb, ()))
        sub_edges = [(s, t) for s, t in raw_edges if s in include and t in include]
        if not sub_edges:
            continue  # nothing to draw (all-orphan schema)
        sub_nodes = [id_node[i] for i in include if i in id_node]
        name = f"Map · {sch}"[:31]
        base = name
        k = 2
        while name.lower() in used_names:
            name = f"{base[:28]}_{k}"
            k += 1
        used_names.add(name.lower())
        _build_lineage_map_sheet(wb, sub_nodes, sub_edges, entity_names, name)
        made += 1


def _build_lineage_map_sheet(wb, all_nodes, raw_edges, entity_names=None, sheet_name="Lineage Map") -> None:
    """Append a lineage-map sheet that mirrors the app: tables AND pipeline/
    job nodes are drawn as boxes, placed in columns by dependency depth
    (sources left → downstream right). Uses the RAW edges (not a collapsed
    table→table graph) so fan-out jobs don't explode into cyclic cross-products.
    Truly disconnected tables are grouped at the bottom. Pure cells — no drawing
    objects — so it's safe to open anywhere.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter

    entity_names = entity_names or {}
    ids = {n.id: n for n in all_nodes}
    if not ids:
        return

    layers_map, orphans, adj_down = _layer_nodes(ids, raw_edges)

    ws = wb.create_sheet(sheet_name[:31])
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="94A3B8")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=12, color=_INDIGO)
    hdr_font = Font(bold=True, size=9, color="64748B")
    arrow_font = Font(bold=True, size=12, color="64748B")

    def is_entity(node) -> bool:
        return getattr(node, "node_type", None) == "entity"

    def box_text(node) -> str:
        if is_entity(node):
            nm = entity_names.get(node.id) or node.display_name or f"{node.entity_type} {node.entity_id[:8]}"
            return f"{nm}\n{node.entity_type}"
        return f"{node.name}\n{node.table_type} · {node.lineage_status}"

    def box_note(node) -> str:
        if is_entity(node):
            nm = entity_names.get(node.id) or node.display_name or node.entity_type
            extra = f"\nlast run: {node.last_run}" if node.last_run else ""
            return f"{nm}\n{node.entity_type} · {node.entity_id}{extra}"
        return (
            f"{node.full_name}\n"
            f"upstream: {node.upstream_count}  ·  downstream: {node.downstream_count}"
            + (f"\nowner: {node.owner}" if node.owner else "")
        )

    def draw_box(r0: int, c0: int, node) -> None:
        # Style every cell in the box BEFORE merging: once merged, the inner
        # cells become read-only MergedCell objects that reject style assignment.
        if is_entity(node):
            fill_hex, font_hex = _ENTITY_FILL, _ENTITY_FONT
        else:
            fill_hex = _STATUS_FILL.get(node.lineage_status, "E5E7EB")
            font_hex = _STATUS_FONT.get(node.lineage_status, "1F2937")
        fill = PatternFill("solid", fgColor=fill_hex)
        for rr in range(r0, r0 + _NODE_H):
            for cc in range(c0, c0 + _NODE_W):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = fill
                cell.border = box_border
        ws.merge_cells(start_row=r0, start_column=c0, end_row=r0 + _NODE_H - 1, end_column=c0 + _NODE_W - 1)
        tl = ws.cell(row=r0, column=c0)
        tl.value = box_text(node)
        tl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        tl.font = Font(bold=True, size=10, color=font_hex)
        c = Comment(box_note(node), "Lineage Explorer")
        c.width, c.height = 320, 90
        tl.comment = c

    # Title + legend
    title = ws.cell(row=_TITLE_ROW, column=1,
                    value="Lineage map — flows left (sources) → right (downstream).  Boxes = tables; indigo boxes = jobs/pipelines.  Hover any box for details.")
    title.font = title_font

    max_layer = max(layers_map) if layers_map else 0
    for L in range(max_layer + 1):
        col0 = 1 + L * (_NODE_W + _GAP_COL)
        h = ws.cell(row=_LAYER_HDR_ROW, column=col0, value=f"LAYER {L}")
        h.font = hdr_font
        for i, nid in enumerate(layers_map.get(L, [])):
            r0 = _FIRST_NODE_ROW + i * (_NODE_H + _GAP_ROW)
            draw_box(r0, col0, ids[nid])
            if adj_down.get(nid) and L < max_layer:
                a = ws.cell(row=r0 + _NODE_H // 2, column=col0 + _NODE_W, value="→")
                a.alignment = Alignment(horizontal="center", vertical="center")
                a.font = arrow_font

    # Truly disconnected nodes grouped below the layered area
    if orphans:
        tallest = max((len(v) for v in layers_map.values()), default=0)
        base = _FIRST_NODE_ROW + tallest * (_NODE_H + _GAP_ROW) + 1
        oh = ws.cell(row=base, column=1, value="NOT CONNECTED — no recorded lineage")
        oh.font = Font(bold=True, size=9, color="92400E")
        per_row = max(max_layer + 1, 4)
        orphans.sort(key=lambda n: (getattr(ids[n], "name", None) or n).lower())
        for k, nid in enumerate(orphans):
            rr = base + 1 + (k // per_row) * (_NODE_H + _GAP_ROW)
            cc = 1 + (k % per_row) * (_NODE_W + _GAP_COL)
            draw_box(rr, cc, ids[nid])

    # Column widths: node columns wide, spacer columns narrow.
    n_orphan_cols = (max(max_layer + 1, 4)) if orphans else 0
    max_cols = max(1 + (max_layer + 1) * (_NODE_W + _GAP_COL), 1 + n_orphan_cols * (_NODE_W + _GAP_COL))
    period = _NODE_W + _GAP_COL
    for col in range(1, max_cols + 1):
        is_spacer = (col - 1) % period == _NODE_W
        ws.column_dimensions[get_column_letter(col)].width = 4 if is_spacer else 9


def build_lineage_workbook(catalog: str, schema: str | None, result, column_edges=None,
                           entity_names=None, table_edges=None) -> bytes:
    """Build a styled .xlsx (bytes) from a LineageResponse.

    entity_names maps entity node id → resolved display name (job/pipeline name).
    table_edges is the list of REAL recorded table→table pairs (from
    get_table_edges): [{source, target, entity_type, entity_id}, ...]. The
    Lineage sheet uses these directly rather than cross-producting a job's
    sources × targets (which fabricates pairs).

    Raises ImportError if openpyxl is unavailable (handled by the caller).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    entity_names = entity_names or {}
    real_edges = table_edges or []

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_INDIGO)
    header_align = Alignment(horizontal="left", vertical="center")
    header_border = Border(bottom=Side(style="thin", color="CBD5E1"))
    title_font = Font(bold=True, size=15, color=_INDIGO)
    label_font = Font(bold=True, color="334155")
    note_font = Font(italic=True, size=10, color="64748B")
    mono_font = Font(name="Consolas", size=10)

    table_nodes = [n for n in result.nodes if getattr(n, "node_type", None) == "table"]
    entity_nodes = [n for n in result.nodes if getattr(n, "node_type", None) == "entity"]
    orphan_count = sum(1 for t in table_nodes if t.lineage_status == "orphan")

    def via_label(etype, eid):
        if not etype or not eid:
            return "—"
        return entity_names.get(f"entity:{etype}:{eid}") or f"{etype} {eid[:8]}"

    def style_header(ws, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"

    def autosize(ws, headers, rows, mins=None, maxs=None):
        mins = mins or {}
        maxs = maxs or {}
        for i, h in enumerate(headers):
            width = len(str(h))
            for r in rows:
                v = r[i]
                if v is not None:
                    width = max(width, len(str(v)))
            width = min(maxs.get(i, 60), max(mins.get(i, 10), width + 2))
            ws.column_dimensions[get_column_letter(i + 1)].width = width

    scope = "schema" if schema else "catalog"
    scope_label = f"{catalog}.{schema}" if schema else catalog

    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Lineage Explorer — export"
    ws["A1"].font = title_font
    summary = [
        ("Scope", scope),
        ("Target", scope_label),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Tables", len(table_nodes)),
        ("Pipelines / entities", len(entity_nodes)),
        ("Lineage edges (table → table)", len(real_edges)),
        ("Column lineage edges", len(column_edges.edges) if column_edges else 0),
        ("Tables without lineage (orphan)", orphan_count),
    ]
    r = 3
    for label, value in summary:
        ws.cell(row=r, column=1, value=label).font = label_font
        if value != "":
            ws.cell(row=r, column=2, value=value)
        if label == "Target":
            ws.cell(row=r, column=2).font = mono_font
        r += 1

    # Notes — explain what lineage does and doesn't cover.
    r += 1
    ws.cell(row=r, column=1, value="Notes").font = label_font
    r += 1
    notes = [
        "Lineage reflects the last 90 days of query-captured activity in Unity Catalog.",
        "Only tables read/written by a tracked query appear connected; others show as 'orphan'.",
        "Edges are the real recorded source→target pairs (with the job/pipeline that ran them).",
        "Column lineage is included for single-schema exports only (empty for catalog scope).",
        "Lineage Map sheets render the flow per schema (sources left → downstream right).",
    ]
    for n in notes:
        ws.cell(row=r, column=1, value="• " + n).font = note_font
        r += 1

    # Legend — status colors used on the Tables sheet and Lineage Map.
    r += 1
    ws.cell(row=r, column=1, value="Legend").font = label_font
    r += 1
    legend = [("root", "source — no upstream"), ("connected", "has upstream & downstream"),
              ("leaf", "sink — no downstream"), ("orphan", "no recorded lineage")]
    for status, desc in legend:
        c = ws.cell(row=r, column=1, value=status)
        c.fill = PatternFill("solid", fgColor=_STATUS_FILL[status])
        c.font = Font(bold=True, color=_STATUS_FONT[status])
        ws.cell(row=r, column=2, value=desc).font = note_font
        r += 1
    c = ws.cell(row=r, column=1, value="pipeline/job")
    c.fill = PatternFill("solid", fgColor=_ENTITY_FILL)
    c.font = Font(bold=True, color=_ENTITY_FONT)
    ws.cell(row=r, column=2, value="a job, pipeline or query that moves data").font = note_font
    r += 2

    # Per-schema breakdown.
    ws.cell(row=r, column=1, value="Per-schema breakdown").font = label_font
    r += 1
    hdr = ws.cell(row=r, column=1, value="Schema"); hdr.font = label_font
    ws.cell(row=r, column=2, value="Tables").font = label_font
    ws.cell(row=r, column=3, value="Orphans").font = label_font
    r += 1
    per_schema = defaultdict(lambda: [0, 0])
    for t in table_nodes:
        _, sch = _split_fqdn(t.full_name)
        per_schema[sch][0] += 1
        if t.lineage_status == "orphan":
            per_schema[sch][1] += 1
    for sch in sorted(per_schema, key=lambda s: -per_schema[s][0]):
        cnt, orph = per_schema[sch]
        ws.cell(row=r, column=1, value=sch).font = mono_font
        ws.cell(row=r, column=2, value=cnt)
        ws.cell(row=r, column=3, value=orph)
        r += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12

    # --- Tables ---
    ws = wb.create_sheet("Tables")
    headers = ["Full name", "Name", "Catalog", "Schema", "Type", "Owner",
               "Upstream", "Downstream", "Status", "Columns", "Comment", "Created", "Updated"]
    ws.append(headers)
    rows = []
    for t in sorted(table_nodes, key=lambda n: n.full_name):
        cat, sch = _split_fqdn(t.full_name)
        rows.append([
            t.full_name, t.name, cat, sch, t.table_type, t.owner or "",
            t.upstream_count, t.downstream_count, t.lineage_status,
            len(t.columns or []), t.comment or "", t.created_at or "", t.updated_at or "",
        ])
    status_col = headers.index("Status")  # 0-based
    for r in rows:
        ws.append(r)
        excel_row = ws.max_row
        ws.cell(row=excel_row, column=1).font = mono_font  # Full name mono
        status = r[status_col]
        if status in _STATUS_FILL:
            cell = ws.cell(row=excel_row, column=status_col + 1)
            cell.fill = PatternFill("solid", fgColor=_STATUS_FILL[status])
            cell.font = Font(bold=True, color=_STATUS_FONT[status])
    style_header(ws, len(headers))
    autosize(ws, headers, rows, mins={0: 30, 10: 20}, maxs={0: 60, 5: 36, 10: 60})

    # --- Lineage (real table → table pairs, with the job/pipeline that ran them) ---
    ws = wb.create_sheet("Lineage")
    headers = ["Source", "Via (pipeline / job)", "Target"]
    ws.append(headers)
    le_rows = sorted(
        ([e["source"], via_label(e.get("entity_type"), e.get("entity_id")), e["target"]] for e in real_edges),
        key=lambda r: (r[0], r[2], r[1]),
    )
    for r in le_rows:
        ws.append(r)
        ws.cell(row=ws.max_row, column=1).font = mono_font
        ws.cell(row=ws.max_row, column=3).font = mono_font
    style_header(ws, len(headers))
    autosize(ws, headers, le_rows, mins={0: 30, 1: 22, 2: 30}, maxs={0: 60, 1: 40, 2: 60})

    # --- Pipelines ---
    if entity_nodes:
        ws = wb.create_sheet("Pipelines")
        headers = ["Type", "Name", "Entity ID", "Last run", "Owner", "Cost (USD, 30d)"]
        ws.append(headers)
        pr_rows = []
        for en in entity_nodes:
            name = en.display_name or entity_names.get(en.id, "")
            pr_rows.append([
                en.entity_type, name, en.entity_id,
                en.last_run or "", en.owner or "", en.cost_usd if en.cost_usd is not None else "",
            ])
        for r in pr_rows:
            ws.append(r)
            ws.cell(row=ws.max_row, column=3).font = mono_font  # Entity ID
            if isinstance(r[5], (int, float)):
                ws.cell(row=ws.max_row, column=6).number_format = '"$"#,##0.00'
        style_header(ws, len(headers))
        autosize(ws, headers, pr_rows, mins={2: 24}, maxs={1: 50, 2: 44})

    # --- Column Lineage ---
    if column_edges and column_edges.edges:
        ws = wb.create_sheet("Column Lineage")
        headers = ["Source table", "Source column", "Target table", "Target column"]
        ws.append(headers)
        cl_rows = [[e.source_table, e.source_column, e.target_table, e.target_column]
                   for e in column_edges.edges]
        for r in cl_rows:
            ws.append(r)
            ws.cell(row=ws.max_row, column=1).font = mono_font
            ws.cell(row=ws.max_row, column=3).font = mono_font
        style_header(ws, len(headers))
        autosize(ws, headers, cl_rows, mins={0: 28, 2: 28}, maxs={0: 54, 1: 36, 2: 54, 3: 36})

    # --- Lineage Map(s): layered colored boxes, left → right flow ---
    # Schema scope → one map. Catalog scope → one map PER SCHEMA (a single
    # catalog-wide grid is dozens of columns wide and interleaves unrelated
    # schemas). Uses ALL nodes (tables + entities) and the RAW edges.
    try:
        raw_edges = [(e.source, e.target) for e in result.edges]
        if schema is not None:
            _build_lineage_map_sheet(wb, result.nodes, raw_edges, entity_names, "Lineage Map")
        else:
            _add_per_schema_maps(wb, result.nodes, raw_edges, entity_names)
    except Exception as map_err:
        # Never let the visual map break the rest of the workbook.
        logger.warning(f"Lineage Map sheet skipped: {map_err}")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
